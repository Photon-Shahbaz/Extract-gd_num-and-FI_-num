from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def extract_gd(driver):

    # Find gd
    time.sleep(5)

    # Create an ActionChains object
    actions = ActionChains(driver)

    for _ in range(9):
        actions.send_keys(Keys.TAB).perform()
        time.sleep(0.5)

    actions.send_keys(Keys.ENTER).perform()

    # Click on gd submitted
    time.sleep(5)

    for _ in range(8):
        actions.send_keys(Keys.ESCAPE).perform()
        time.sleep(0.25)
        actions.send_keys(Keys.TAB).perform()
    time.sleep(0.5)

    # actions.send_keys(Keys.TAB).perform()
    # actions.key_down(Keys.SHIFT).send_keys(Keys.TAB).key_up(Keys.SHIFT).perform()

    actions.send_keys(Keys.ENTER).perform()
    # print("tab key")

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import os

def extract_rows(driver):

    # Switch to iframe
    WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.ID, "frame"))
    )

    # Wait until table appears
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder2_dgSubmitted"))
    )

    extracted_data = []

    table_xpath = '//*[@id="ctl00_ContentPlaceHolder2_dgSubmitted"]/tbody'

    # Extract first 5 rows (tr[2] to tr[6])
    for i in range(2, 7):
        row_xpath = f"{table_xpath}/tr[{i}]"

        col1 = driver.find_element(By.XPATH, row_xpath + '/td[1]').get_attribute("innerText").strip()
        col2 = driver.find_element(By.XPATH, row_xpath + '/td[2]').get_attribute("innerText").strip()
        col3 = driver.find_element(By.XPATH, row_xpath + '/td[3]').get_attribute("innerText").strip()

        col5_element = driver.find_element(By.XPATH, row_xpath + '/td[5]//a')
        col5 = col5_element.get_attribute("innerText").strip()

        extracted_data.append([col1, col2, col3, col5])

    driver.switch_to.default_content()

    # Excel file handling
    file_name = "PSW_updates.xlsx"

    # Load workbook or create new
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active

        # Read existing Column1 values to avoid duplicates
        existing_keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing_keys.add(str(row[0]).strip())

    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Column1", "Column2", "Column3", "Column5"])
        existing_keys = set()

    # Append only new entries
    new_entries_count = 0
    new_entries_list = []  # To store only newly added rows for printing

    for row in extracted_data:
        if row[0] not in existing_keys:  # Check Column1 for uniqueness
            ws.append(row)
            existing_keys.add(row[0])
            new_entries_count += 1
            new_entries_list.append(row)

    wb.save(file_name)

    print(f"\n✅ New unique records added in this run: {new_entries_count}")

    # Print new entries details in console
    if new_entries_count > 0:
        print("\n📌 Newly added entries:")
        for entry in new_entries_list:
            print(f"  • {entry}")
    else:
        print("ℹ️ No new entries. All values were already present.")

    print(f"\n📂 File saved/updated: {file_name}\n")



    return extracted_data
